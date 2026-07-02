from app.reports.context import ReportContext
from app.reports.renderers.excel_renderer import render_excel
from app.reports.renderers.html_renderer import render_html
from app.reports.renderers.pdf_renderer import render_pdf


def _context(**overrides) -> ReportContext:
    defaults = dict(
        audit_job_id="job-1",
        tenant_name="Acme Corp",
        provider="aws",
        job_status="completed",
        created_at="2026-01-01T00:00:00Z",
        findings=[
            {
                "id": "f-1",
                "module": "validate",
                "finding_type": "unauthorized_peering",
                "severity": "high",
                "title": "Unauthorized VPC peering",
                "description": "Details here.",
                "status": "open",
            }
        ],
        severity_counts={"high": 1},
        exec_summary="Overall risk is moderate.",
    )
    defaults.update(overrides)
    return ReportContext(**defaults)


def test_render_html_includes_tenant_and_findings():
    html = render_html(_context(), "technical")
    assert "Acme Corp" in html
    assert "Unauthorized VPC peering" in html
    assert "Overall risk is moderate." in html


def test_render_html_executive_omits_detailed_findings_table():
    html = render_html(_context(), "executive")
    assert "Detailed Findings" not in html


def test_render_pdf_produces_nonempty_bytes():
    pdf_bytes = render_pdf(_context(), "technical")
    assert isinstance(pdf_bytes, bytes)
    assert pdf_bytes.startswith(b"%PDF")


def test_render_pdf_handles_no_findings():
    pdf_bytes = render_pdf(_context(findings=[], severity_counts={}), "compliance")
    assert pdf_bytes.startswith(b"%PDF")


def test_render_excel_produces_valid_workbook():
    from io import BytesIO

    from openpyxl import load_workbook

    excel_bytes = render_excel(_context(), "compliance")
    workbook = load_workbook(BytesIO(excel_bytes))
    assert "Summary" in workbook.sheetnames
    assert "Findings" in workbook.sheetnames
    findings_sheet = workbook["Findings"]
    assert findings_sheet.cell(row=2, column=4).value == "Unauthorized VPC peering"
