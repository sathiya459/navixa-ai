import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.reports.context import ReportContext

HEADER_FILL = PatternFill(start_color="0B3D91", end_color="0B3D91", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def render_excel(context: ReportContext, report_type: str) -> bytes:
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["NAVIXA AI", f"{report_type.capitalize()} Report"])
    summary_sheet.append(["Tenant", context.tenant_name])
    summary_sheet.append(["Provider", context.provider.upper()])
    summary_sheet.append(["Audit Job", context.audit_job_id])
    summary_sheet.append(["Status", context.job_status])
    summary_sheet.append(["Audit Created", context.created_at])
    summary_sheet.append([])
    summary_sheet.append(["Severity", "Count"])
    for severity, count in context.severity_counts.items():
        summary_sheet.append([severity, count])

    if context.exec_summary:
        summary_sheet.append([])
        summary_sheet.append(["Executive Summary"])
        summary_sheet.append([context.exec_summary])

    findings_sheet = workbook.create_sheet("Findings")
    headers = ["Module", "Finding Type", "Severity", "Title", "Description", "Status"]
    findings_sheet.append(headers)
    for cell in findings_sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    for finding in context.findings:
        findings_sheet.append(
            [
                finding["module"],
                finding["finding_type"],
                finding["severity"],
                finding["title"],
                finding["description"],
                finding["status"],
            ]
        )

    for column_cells in findings_sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells if cell.value) if column_cells else 10
        findings_sheet.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 60)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
